import openpyxl
from openpyxl.styles import PatternFill
from ortools.sat.python import cp_model

# 模型初始化
model = cp_model.CpModel()

# 定義參數
month = 9
people_list = ['小明', '小花', '小白', '小城', '小胖']
num_people = len(people_list)
num_days = 25  # 計劃的天數

# 每天最多可休息人數列表
max_rest_per_day = [
    2,3,1,3,1,3,1,1,1,2,1,1,3,2,2,3,3,3,2,3,1,2,2,3,1
]

# 固定休假日
mandatory_off = [
    [1, 6, 14],  # 小明的固定休假日
    [4, 16, 20], # 小花的固定休假日
    [2, 10, 18], # 小白的固定休假日
    [5, 15, 21], # 小城的固定休假日
    [3, 12, 19]  # 小胖的固定休假日
]

# 定義變量：work[p][d] 表示第 p 個人在第 d 天是否上班 (1 表示上班，0 表示休息)
work = []
for p in range(num_people):
    person_schedule = []
    for d in range(num_days):
        person_schedule.append(model.NewBoolVar(f'work_{p}_{d}'))
    work.append(person_schedule)

# 約束條件 1：每個人休息 10 天（包括固定休假）
for p in range(num_people):
    model.Add(sum(1 - work[p][d] for d in range(num_days)) == 10)

# 約束條件 2：連續上班不能超過 4 天
for p in range(num_people):
    for d in range(num_days - 4):
        model.Add(sum(work[p][d + i] for i in range(5)) <= 4)

# 約束條件 3：不可單獨上班一天，第一天和最後一天例外
for p in range(num_people):
    for d in range(1, num_days - 1):
        model.AddBoolOr([
            work[p][d - 1],
            work[p][d + 1]
        ]).OnlyEnforceIf(work[p][d])

# 約束條件 4：固定休假日
for p in range(num_people):
    for day in mandatory_off[p]:
        model.Add(work[p][day - 1] == 0)

# 新增的約束條件 5：每天最多可休息的人數，使用 max_rest_per_day 列表
for d in range(num_days):
    model.Add(sum(1 - work[p][d] for p in range(num_people)) <= max_rest_per_day[d])

# 創建求解器
solver = cp_model.CpSolver()

# 設定求解器參數
solver.parameters.max_time_in_seconds = 60

# 求解
status = solver.Solve(model)

# 構建 Excel 數據
if status == cp_model.FEASIBLE or status == cp_model.OPTIMAL:
    # 創建一個新的 Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "排班表"

    # 定義填充樣式
    fill_mandatory_off = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黃色底色（指休）
    fill_vacation = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 綠色底色（休假）

    # 第一行顯示日期
    ws.append(['日期'] + [f'{month}/{d + 1}' for d in range(num_days)])

    # 添加當日可休人數
    ws.append(['當日可休人力'] + max_rest_per_day)

    # 輸出每個人的排班狀態
    for p in range(num_people):
        row = [people_list[p]]
        for d in range(num_days):
            if d + 1 in mandatory_off[p]:
                row.append('指休')
            elif solver.Value(work[p][d]) == 0:
                row.append('休假')
            else:
                row.append('上班')
        ws.append(row)

    # 設置單元格的背景顏色
    for p in range(num_people):
        for d in range(num_days):
            cell = ws.cell(row=p + 3, column=d + 2)  # +3 是因為要跳過標題行和當日可休人數行
            if d + 1 in mandatory_off[p]:
                cell.fill = fill_mandatory_off  # 指休
            elif solver.Value(work[p][d]) == 0:
                cell.fill = fill_vacation  # 休假

    # 保存 Excel 文件
    excel_file = './schedule_output.xlsx'
    wb.save(excel_file)

    # 文件路徑返回
    excel_file
else:
    "No solution found within the time limit."
